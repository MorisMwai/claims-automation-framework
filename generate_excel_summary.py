import pandas as pd
import os
import json
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList

# ---------------------------
# UTILITY: Format Worksheet
# ---------------------------

def format_sheet(ws):
    """
    Apply formatting to worksheet:
    - Bold headers
    - Center alignment
    - Bottom borders
    - Auto-adjust column width
    """
    header_font = Font(bold=True)
    align_center = Alignment(horizontal='center')
    border = Border(bottom=Side(border_style='thin'))

    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = align_center
        cell.border = border

    for column_cells in ws.columns:
        max_len = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = max_len + 4

# ---------------------------
# MAIN REPORT LOGIC
# ---------------------------

def generate_report(json_total, json_valid, json_invalid, json_type, json_branch, json_avg, output_path):
    # Convert incoming DataFrames
    # Convert JSON strings to DataFrames
    total_df = pd.DataFrame(json.loads(json_total))
    valid_df = pd.DataFrame(json.loads(json_valid))
    invalid_df = pd.DataFrame(json.loads(json_invalid))
    type_df = pd.DataFrame(json.loads(json_type))
    branch_df = pd.DataFrame(json.loads(json_branch))
    avg_df = pd.DataFrame(json.loads(json_avg))
    
    # Create a new workbook and remove default sheet
    wb = Workbook()
    wb.remove(wb.active)

    # 1️ SUMMARY SHEET
    ws_summary = wb.create_sheet("Summary")

    # Prepare summary metrics
    summary_data = {
        "Metric": ["Total Claims", "Valid Claims", "Invalid Claims", "Avg Claim Amount (Valid)"],
        "Value": [
            int(total_df.iloc[0, 0]),
            int(valid_df.iloc[0, 0]),
            int(invalid_df.iloc[0, 0]),
            round(float(avg_df.iloc[0, 0]), 2)
        ]
    }
    summary_df = pd.DataFrame(summary_data)

    # Append summary data
    for row in dataframe_to_rows(summary_df, index=False, header=True):
        ws_summary.append(row)

    # Format the worksheet
    format_sheet(ws_summary)

    # Format average claim amount as KES currency and auto-fit column
    for row in ws_summary.iter_rows(min_row=5, max_row=5, min_col=2, max_col=2):
        for cell in row:
            cell.number_format = '"KES" #,##0.00'
            ws_summary.column_dimensions[cell.column_letter].width = 20  

    # Add bar chart comparing Valid/Invalid/Total claims
    chart = BarChart()
    chart.title = "Valid vs Invalid Claims"
    chart.y_axis.title = "Number of Claims"
    chart.x_axis.title = "Claim Category"

    data = Reference(ws_summary, min_col=2, min_row=2, max_row=4)  # rows 2 to 4
    categories = Reference(ws_summary, min_col=1, min_row=2, max_row=4)
    chart.add_data(data, titles_from_data=False)
    chart.set_categories(categories)
    chart.style = 5
    chart.height = 8
    chart.width = 16

    # Add labels to bars
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showVal = True
    chart.dataLabels.showCatName = False
    chart.dataLabels.showSerName = False

    ws_summary.add_chart(chart, "D2")

    # 2️ BY_CLAIM_TYPE SHEET
    ws_type = wb.create_sheet("By_Claim_Type")
    for row in dataframe_to_rows(type_df, index=False, header=True):
        ws_type.append(row)
    format_sheet(ws_type)

    # 3️ BY_BRANCH SHEET
    ws_branch = wb.create_sheet("By_Branch")
    for row in dataframe_to_rows(branch_df, index=False, header=True):
        ws_branch.append(row)
    format_sheet(ws_branch)

    # Save the final workbook
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    print(f"Summary report saved to: {output_path}")
    
# ---------------------------
# ENTRY POINT FOR UiPATH
# ---------------------------

def main(json_total, json_valid, json_invalid, json_type, json_branch, json_avg, output_path):
    generate_report(json_total, json_valid, json_invalid, json_type, json_branch, json_avg, output_path)

# ---------------------------
# STANDALONE TEST HARNESS
# ---------------------------

# if __name__ == "__main__":
#     generate_report(json_total, json_valid, json_invalid, json_type, json_branch, json_avg, output_path)
    # Mock JSON strings
    # import json

    # data = {
    #     "total": [{"TotalClaims": 100}],
    #     "valid": [{"ValidClaims": 70}],
    #     "invalid": [{"InvalidClaims": 30}],
    #     "avg": [{"Average_Claim_Amount": 402837.89}],
    #     "type": [
    #         {"Claim_Type": "Health", "Count": 24},
    #         {"Claim_Type": "Motor", "Count": 28},
    #         {"Claim_Type": "Property", "Count": 30},
    #         {"Claim_Type": "Education", "Count": 18}
    #     ],
    #     "branch": [
    #         {"Branch": "Kenya", "Count": 23},
    #         {"Branch": "Uganda", "Count": 20},
    #         {"Branch": "South Sudan", "Count": 27},
    #         {"Branch": "Malawi", "Count": 30}
    #     ]
    # }

    # generate_report(
    #     json.dumps(data["total"]),
    #     json.dumps(data["valid"]),
    #     json.dumps(data["invalid"]),
    #     json.dumps(data["type"]),
    #     json.dumps(data["branch"]),
    #     json.dumps(data["avg"]),
    #     r"C:\Users\MorisMwaiWachira\Desktop\MorisMwai_RPA_Assignment\CIC_Claims_Automation_Assessment\Data\Output\claims_summary.xlsx"
    # )